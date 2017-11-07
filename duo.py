# -*- coding: utf-8 -*-

#import the modules needed for this to work
import datetime
import time
import duo_client

#Set the time range to get from the Duo logs
#This gets the past 24 hours logs
timerange = int(time.time()) - 86400

#import the Python Excel Module and create a workbook object
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active

#Name the sheet we are writing to
ws_sheet = wb.get_sheet_by_name('Sheet')
ws_sheet.title = 'Duo Log In Attempts'
    
#Set the column titles and widths for the workbook
a = ws['A1']
a.font = Font(bold=True)
ws['A1'] = 'Username'
ws.column_dimensions["A"].width = 12

b = ws['B1']
b.font = Font(bold=True)
ws['B1'] = 'Auth Method'
ws.column_dimensions["B"].width = 22

c = ws['C1']
c.font = Font(bold=True)
ws['C1'] = 'Device'
ws.column_dimensions["C"].width = 30

d = ws['D1']
d.font = Font(bold=True)
ws['D1'] = 'Result'
ws.column_dimensions["D"].width = 10

e = ws['E1']
e.font = Font(bold=True)
ws['E1'] = 'Reason'
ws.column_dimensions["E"].width = 20

f = ws['F1']
f.font = Font(bold=True)
ws['F1'] = 'Time'
ws.column_dimensions["F"].width = 25

g = ws['G1']
g.font = Font(bold=True)
ws['G1'] = 'Integration'
ws.column_dimensions["F"].width = 25

#Authtenticate to Duo with our API creds and server
admin_api = duo_client.Admin(
    ikey=('YOUR_IKEY'),
    skey=('YOUR_SKEY'),
    host=('YOUR.HOST.duosecurity.com'),
)

#Create a variable for the logs
#Call theAPI and get the logs for the past 24 hours using the timerange varaible created earlier
logs = admin_api.get_authentication_log(timerange)

#Loop through and get the vaules for the report

for i in range(len(logs)):
    user = (logs[i]['username'])
    authmethod = logs[i]['factor']
    authdevice = logs[i]['device']
    authenticated = logs[i]['result']
    why = logs[i]['reason']
    where = logs[i]['integration']
    
#Since the logs are in Unix time stamp format we have to convert them to something understandable
    authtime = logs[i]['timestamp']
    converted_time = datetime.datetime.fromtimestamp(authtime)
    timeformat = converted_time.strftime('%Y-%m-%d %H:%M:%S')
    
#Write the requested values to the spreadsheet report    
    ws.append([user, authmethod, authdevice, authenticated, why, timeformat, where])
    
#save the report
    wb.save("duo.xlsx")

    
       
    
    #This is a test line to see what values we get back
    #This is useful for debugging the code
    #print (user, authmethod, authdevice, authenticated, why, (converted_time.strftime('%Y-%m-%d %H:%M:%S')))
